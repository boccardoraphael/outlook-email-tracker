"""
email_tracker.py

Reads client folders from the Terra Translations Outlook mailbox,
extracts structured data, and writes a clean tracker Excel file on OneDrive.

Usage:
    py email_tracker.py

Requires: pywin32, openpyxl, langdetect
"""

import sys
import re
from pathlib import Path

try:
    import win32com.client
except ImportError:
    sys.exit("Missing dependency: run  py -m pip install pywin32")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: run  py -m pip install openpyxl")

try:
    from langdetect import detect, LangDetectException
except ImportError:
    sys.exit("Missing dependency: run  py -m pip install langdetect")


# ── Configuration ─────────────────────────────────────────────────────────────

EXCEL_PATH = Path(r"C:\Users\DESKTOP\OneDrive - Terra Translations\Documents\Client Email Tracker.xlsx")

QUOTED_PMS = {
    "gustavo teixeira", "bruno teixeira", "estéfano vitagliano",
    "julia sestari", "dieggo lessa", "alex alves",
    "mateus andrade", "rodrigo campos", "thiago kunis", "samuel coelho",
}

SKIP_SENDER_DOMAINS  = {"wizer-training.com", "microsoft.com", "office365.com", "sharepoint.com"}
SKIP_SENDER_EMAILS   = {"no-reply@wizer-training.com", "terrasoft@terrateamup.com",
                        "memoq@terrateamup.com", "noreply@terrateamup.com"}
SKIP_SENDER_KEYWORDS = {"onboarding", "security awareness", "microsoft teams"}
RAPHAEL_EMAIL        = "raphael.boccardo@terrateamup.com"

# ── Project relevance filter ───────────────────────────────────────────────────

# Task types that are always project-related
PROJECT_TASK_TYPES = {
    "Translation", "Proofreading", "Review", "LQA",
    "Post-Editing", "Delivery", "Handoff", "Bug Report",
    "Timing Fix", "Credits", "Query",
}

# Subject/body keywords that confirm project relevance
_PROJECT_SIGNAL = re.compile(
    r"\b(translation|proofreading|proofread|delivery|delivered|entregou|"
    r"batch|lote|handoff|hand-off|\bho\b|bug report|timing|credits|"
    r"word count|wordcount|deadline|prazo|linguist|linguística|"
    r"review|lqa|post.edit|locali[sz]ation|locali[sz]ação|"
    r"project|projeto|segment|string|file|arquivo|"
    r"aspera|memoq|smartcat|phrase|trados|studio|"
    r"\b[A-Z]{2,6}[_\-]\w+\b)\b",  # batch/file codes like TL_batch_33
    re.IGNORECASE,
)

# Keywords that confirm NON-project (HR/P&C/admin/chat)
_NON_PROJECT_SIGNAL = re.compile(
    r"\b(people.?(and|&).?culture|human.?resources?\b|hr\b|onboard|"
    r"boas.?vindas|bem.?vindo|welcome.?to.?terra|"
    r"birthday|aniversário|anniversary|"
    r"salary|salário|payroll|folha|benefíc|benefit|"
    r"vacation|férias|time.?off|folga|"
    r"policy|política|compliance|code.?of.?conduct|"
    r"performance.?review|avaliação.?de.?desempenho|"
    r"you.?have.?been.?added|added.?you.?to|"
    r"team.?notification|missed.?activity|"
    r"open.?assignment|training.?course|"
    r"invoice|nota.?fiscal|payment|pagamento|fatura)\b",
    re.IGNORECASE,
)


def is_project_related(subject: str, body: str, task_type: str) -> bool:
    """Return True only if this email is clearly about a project/task."""
    # Always keep if task type is project-related
    if task_type in PROJECT_TASK_TYPES:
        return True

    # Always keep Quoted delivery pattern: "MM-DD-HHhMM - [CLIENT] - ..."
    if _QUOTED_SUBJ.match(re.sub(r"^(RE|FW|FWD):\s*", "", subject, flags=re.IGNORECASE)):
        return True

    combined = (subject + " " + body[:800]).lower()

    # Hard exclude if it matches non-project patterns
    if _NON_PROJECT_SIGNAL.search(combined):
        return False

    # Keep if it has project signals
    if _PROJECT_SIGNAL.search(combined):
        return True

    # Also keep if subject has a project code like [i38] or [TL_batch]
    if re.search(r"\[[A-Z0-9][A-Z0-9_\- ]{1,24}\]", subject, re.IGNORECASE):
        return True

    return False

SKIP_FOLDERS = {
    "epic smartling",
    "itens enviados", "sent items", "itens excluídos", "deleted items",
    "lixo eletrônico", "junk email", "junk", "rascunhos", "drafts",
    "caixa de saída", "outbox", "calendário", "calendar",
    "contatos", "contacts", "tarefas", "tasks", "notas", "notes",
    "observações", "arquivo morto", "archive", "diário", "journal",
    "feeds rss", "rss feeds", "problemas de sincronização", "sync issues",
    "conflitos", "falhas locais", "falhas do servidor",
    "histórico de conversa", "conversation history",
    "raiz do yammer", "yammer root", "social activity notifications",
    "quick step settings", "configurações de etapa rápida",
    "conversation action settings", "externalcontacts", "personmetadata",
    "files", "aniversários", "anniversaries", "feriados de brasil",
}

# ── Columns ────────────────────────────────────────────────────────────────────

COLUMNS = [
    "Date Received",    # A
    "Client",           # B
    "Project",          # C  — parsed
    "PM Responsible",   # D
    "From",             # E
    "Task Type",        # F
    "Provider",         # G
    "Deadline",         # H
    "Word Count",       # I  — yellow if missing on translation row
    "Reply/FW",         # J  — RE / FW / blank
    "Project Code",     # K  — [i38], [EDMM], [Thallium]
    "Batch / File",     # L  — TL_batch_33, filename
    "Subject Topic",    # M  — the human-readable part
    "Summary",          # N  — AI-style local summary in English
    "Entry ID",         # O  — hidden dedup
]

# ── Styles ─────────────────────────────────────────────────────────────────────

HEADER_FILL     = PatternFill("solid", fgColor="1F3864")
HEADER_FONT     = Font(color="FFFFFF", bold=True, size=10)
FILL_TRANS_A    = PatternFill("solid", fgColor="D6E4F7")
FILL_TRANS_B    = PatternFill("solid", fgColor="BDD7EE")
FILL_GEN_A      = PatternFill("solid", fgColor="E2EFDA")
FILL_GEN_B      = PatternFill("solid", fgColor="C6EFCE")
FILL_WC_MISSING = PatternFill("solid", fgColor="FFEB9C")
FONT_WC_MISSING = Font(color="9C6500", italic=True, size=9)
BORDER          = Border(bottom=Side(style="thin", color="B0B0B0"))

COLUMN_WIDTHS = {
    "Date Received": 13, "Client": 14, "Project": 20, "PM Responsible": 20,
    "From": 26, "Task Type": 15, "Provider": 22, "Deadline": 15,
    "Word Count": 11, "Reply/FW": 9, "Project Code": 14,
    "Batch / File": 28, "Subject Topic": 35, "Summary": 60, "Entry ID": 5,
}

# ── Text helpers ───────────────────────────────────────────────────────────────

def safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)
    return s


def clean_body(body: str) -> str:
    body = re.sub(r"Geralmente.*?importante\s*<[^>]+>", "", body, flags=re.DOTALL)
    body = re.sub(r"<https?://\S+>", "", body)
    body = re.sub(r"https?://\S+", "", body)
    body = re.sub(r"[ \t]{2,}", " ", body)
    return body.strip()


def detect_lang(text: str) -> str:
    if not text.strip():
        return "unknown"
    try:
        return detect(text[:600])
    except LangDetectException:
        return "unknown"


# ── Subject splitter ───────────────────────────────────────────────────────────

_REPLY_FW   = re.compile(r"^((?:RE|FW|FWD|ENC|RES|R|Rif):\s*)+", re.IGNORECASE)
_PROJ_CODE  = re.compile(r"\[([A-Z0-9][A-Z0-9_\- ]{1,24})\]", re.IGNORECASE)
_NP_SUBJ    = re.compile(
    r"(?P<date>\d{2}-\d{2})-(?P<time>\d{2}h\d{2})"
    r".*?\[(?P<provider>[^\]]+?)\s*-\s*(?P<task>Translation|Proofreading|Review|LQA|Post-Editing)\]",
    re.IGNORECASE,
)
_NP_PROJECT = re.compile(r"\[Native Prime \| (?P<proj>[^\]]+?)\s*(?:\([^)]+\))?\]", re.IGNORECASE)

# Quoted general delivery: "03-20-10h00 - [EA] - 4629909 Metadata Star Wars The Old [111 w]"
_QUOTED_SUBJ = re.compile(
    r"(?P<month>\d{2})-(?P<day>\d{2})-(?P<time>\d{2}h\d{2})"
    r"\s*-\s*\[(?P<client>[^\]]+)\]"
    r"\s*-\s*(?P<rest>.+)$",
    re.IGNORECASE,
)
_WC_BRACKET   = re.compile(r"\[(?P<wc>\d+)\s*w(?:ords?)?\]", re.IGNORECASE)
_TASK_BRACKET = re.compile(
    r"\[(?P<provider>[^\]]+?)\s*-\s*(?P<task>Translation|Proofreading|Review|LQA|Post-Editing)\]",
    re.IGNORECASE,
)

_BATCH      = re.compile(r"\b([A-Z]{1,6}[_\-]\w{2,}(?:[_\-]\w+)*)\b")
_FILENAME   = re.compile(r"\b(\w{4,}\.\w{2,4})\b")


def split_subject(raw: str):
    """Return (reply_fw, project_code, batch_file, topic)."""
    s = raw.strip()

    # Reply/FW prefix
    reply_fw = ""
    m = _REPLY_FW.match(s)
    if m:
        # normalise to last type: RE or FW
        prefix = m.group(0)
        reply_fw = "FW" if re.search(r"\bFW", prefix, re.IGNORECASE) else "RE"
        s = s[m.end():].strip()

    # Native Prime structured subject — extract parts cleanly
    np = _NP_SUBJ.search(s)
    if np:
        np_proj = _NP_PROJECT.search(s)
        project_code = np_proj.group("proj").strip() if np_proj else "Native Prime"
        file_match = re.search(r"\|\s*\[.*?\]\s*(.*?)\s*\[", s)
        batch_file = file_match.group(1).strip() if file_match else ""
        topic = f"{np.group('task')} delivery by {np.group('provider').strip()}"
        return reply_fw, project_code, batch_file, topic

    # Quoted general delivery: "03-20-10h00 - [EA] - 4629909 Project Name [111 w]"
    qm = _QUOTED_SUBJ.match(s)
    if qm:
        project_code = qm.group("client").strip()
        rest = qm.group("rest").strip()
        # Extract word count bracket
        wc_m = _WC_BRACKET.search(rest)
        rest_clean = _WC_BRACKET.sub("", rest).strip()
        # Extract task/provider bracket
        task_m = _TASK_BRACKET.search(rest_clean)
        rest_clean = _TASK_BRACKET.sub("", rest_clean).strip() if task_m else rest_clean
        # Job ID (leading number) goes into batch_file
        job_id_m = re.match(r"^(\d{5,})\s*", rest_clean)
        if job_id_m:
            batch_file = job_id_m.group(1)
            topic = rest_clean[job_id_m.end():].strip(" -")
        else:
            batch_file = ""
            topic = rest_clean.strip(" -")
        return reply_fw, project_code, batch_file, topic

    # Project code(s) in brackets
    codes = _PROJ_CODE.findall(s)
    # filter out known noise words
    noise = {"native prime", "kiln", "hawaii west", "thallium"}
    codes = [c for c in codes if c.lower() not in noise]
    project_code = codes[0] if codes else ""

    # Remove brackets from remaining string
    remainder = re.sub(r"\[[^\]]*\]", "", s).strip(" -:")

    # Batch/file name
    batch_file = ""
    bm = _BATCH.search(remainder)
    if bm:
        batch_file = bm.group(1)
        remainder = remainder.replace(batch_file, "").strip(" -:")
    elif _FILENAME.search(remainder):
        fm = _FILENAME.search(remainder)
        batch_file = fm.group(1)
        remainder = remainder.replace(batch_file, "").strip(" -:")

    topic = re.sub(r"\s{2,}", " ", remainder).strip(" -:")

    return reply_fw, project_code, batch_file, topic


# ── Parsers ────────────────────────────────────────────────────────────────────

_DEADLINE_PT  = re.compile(r"pra\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})")
_DEADLINE_EN  = re.compile(
    r"by\s+((?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
    r"\s+\d{1,2}(?:st|nd|rd|th)?(?:,?\s+\d{4})?)", re.IGNORECASE)
_DEADLINE_ISO = re.compile(r"\b(\d{4}-\d{2}-\d{2})\b")

_WC_PATTERNS = [
    re.compile(r"(\d[\d,\.]+)\s*(?:words?|wds?)\b", re.IGNORECASE),
    re.compile(r"word\s*count[:\s]+(\d[\d,\.]+)", re.IGNORECASE),
    re.compile(r"total[:\s]+(\d[\d,\.]+)\s*words?", re.IGNORECASE),
]

_PM_GREETING = re.compile(
    r"(?:Ol[aá]|Hi|Hello|Dear)[,\s]+([A-ZÀ-Ö][a-zà-ö]+(?:\s+[A-ZÀ-Ö][a-zà-ö]+)?)",
)

_TASK_KEYWORDS = {
    "translation": "Translation", "proofreading": "Proofreading",
    "proofread": "Proofreading", "review": "Review", "lqa": "LQA",
    "post-editing": "Post-Editing", "post editing": "Post-Editing",
    "delivery": "Delivery", "entregou": "Delivery", "delivered": "Delivery",
    "bug report": "Bug Report", "timing": "Timing Fix", "credits": "Credits",
    "handoff": "Handoff", "hand-off": "Handoff", "ho": "Handoff",
    "query": "Query", "question": "Query",
}


def parse_metadata(subject: str, body: str) -> dict:
    r = {"Project": "", "Task Type": "", "Provider": "", "Deadline": "", "Word Count": ""}

    # Native Prime structured
    np = _NP_SUBJ.search(subject)
    if np:
        r["Provider"]  = np.group("provider").strip()
        r["Task Type"] = np.group("task").capitalize()
        r["Deadline"]  = f"{np.group('date')} {np.group('time').replace('h', ':')}"
    np_proj = _NP_PROJECT.search(subject)
    if np_proj:
        r["Project"] = np_proj.group("proj").strip()

    # Quoted general delivery: "03-20-10h00 - [EA] - 4629909 Star Wars [111 w]"
    if not r["Deadline"]:
        qm = _QUOTED_SUBJ.match(subject)
        if qm:
            r["Deadline"] = f"{qm.group('month')}-{qm.group('day')} {qm.group('time').replace('h', ':')}"
            if not r["Project"]:
                # project name is the human text after the job ID
                rest = qm.group("rest")
                rest = _WC_BRACKET.sub("", rest)
                rest = _TASK_BRACKET.sub("", rest)
                rest = re.sub(r"^\d{5,}\s*", "", rest).strip(" -")
                r["Project"] = rest[:40] if rest else qm.group("client")
            # word count from bracket
            if not r["Word Count"]:
                wc_m = _WC_BRACKET.search(qm.group("rest"))
                if wc_m:
                    r["Word Count"] = wc_m.group("wc")
            # task type from bracket
            if not r["Task Type"]:
                task_m = _TASK_BRACKET.search(subject)
                if task_m:
                    r["Provider"]  = task_m.group("provider").strip()
                    r["Task Type"] = task_m.group("task").capitalize()

    # Generic project code fallback
    if not r["Project"]:
        codes = _PROJ_CODE.findall(subject)
        noise = {"native prime", "kiln", "hawaii west", "thallium"}
        codes = [c for c in codes if c.lower() not in noise]
        if codes:
            r["Project"] = codes[0]

    # Deadline from body
    if not r["Deadline"]:
        dl = _DEADLINE_PT.search(body)
        if dl:
            r["Deadline"] = f"{dl.group(1)} {dl.group(2)}"
        else:
            dl = _DEADLINE_EN.search(body) or _DEADLINE_ISO.search(body)
            if dl:
                r["Deadline"] = dl.group(1)

    # Word count from body
    if not r["Word Count"]:
        for pat in _WC_PATTERNS:
            wc = pat.search(body)
            if wc:
                r["Word Count"] = wc.group(1).replace(",", "").replace(".", "")
                break

    # Task type
    if not r["Task Type"]:
        combined = (subject + " " + body[:600]).lower()
        for kw, label in _TASK_KEYWORDS.items():
            if kw in combined:
                r["Task Type"] = label
                break
        if not r["Task Type"]:
            r["Task Type"] = "General"

    return r


def parse_pm(body: str, hint: str) -> str:
    m = _PM_GREETING.search(body)
    if m:
        name = m.group(1).strip()
        if name.lower() not in {"all", "team", "everyone", "there", "you"}:
            return name
    hint_lower = hint.lower()
    for pm in QUOTED_PMS:
        if pm in hint_lower:
            return pm.title()
    return ""


# ── Local summarizer ───────────────────────────────────────────────────────────

# PT→EN key phrase translations for common patterns
_PT_PATTERNS = [
    (re.compile(r"entregou\s+(Translation|Proofreading|Review)", re.I),
     lambda m: f"Delivered {m.group(1)}"),
    (re.compile(r"pra\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})", re.I),
     lambda m: f"Deadline: {m.group(1)} at {m.group(2)}"),
    (re.compile(r"projeto\s+(P-\w+)", re.I),
     lambda m: f"Project {m.group(1)}"),
    (re.compile(r"Work group[:\s]", re.I),
     lambda m: "Work group link included."),
]

# Sentences to skip in summary
_SKIP_SUMMARY = re.compile(
    r"(geralmente|saiba por que|this is an automated|you received this|"
    r"unsubscribe|click here|view in browser|privacy policy|"
    r"microsoft teams|you have been added|bem-vindo|welcome to terra)",
    re.IGNORECASE,
)

# Action verbs that signal importance
_ACTION_SIGNAL = re.compile(
    r"\b(please|could you|can you|kindly|urgent|asap|deadline|deliver|"
    r"send|attached|review|check|confirm|update|fix|bug|issue|"
    r"entregou|prazo|urgente|favor|enviar|verificar|confirmar|corrigir)\b",
    re.IGNORECASE,
)


def summarize(body: str, task_type: str, provider: str, deadline: str) -> str:
    """Build a concise English summary from the email body."""
    if not body.strip():
        return ""

    lang = detect_lang(body[:500])

    # For structured Native Prime delivery notifications, build directly
    if task_type in {"Translation", "Proofreading", "Review", "Delivery"} and provider:
        parts = [f"{task_type} delivered by {provider}."]
        if deadline:
            parts.append(f"Deadline: {deadline}.")
        # Check for project ref
        proj_m = re.search(r"projeto\s+(P-\w+)", body, re.IGNORECASE)
        if proj_m:
            parts.append(f"Project: {proj_m.group(1)}.")
        return " ".join(parts)

    # For PT bodies, apply pattern translations first
    translated_hints = []
    if lang == "pt":
        for pat, fn in _PT_PATTERNS:
            m = pat.search(body)
            if m:
                translated_hints.append(fn(m))

    # Split into sentences and score them
    sentences = re.split(r"(?<=[.!?])\s+|\n", body)
    scored = []
    for sent in sentences:
        sent = sent.strip()
        if len(sent) < 15 or len(sent) > 300:
            continue
        if _SKIP_SUMMARY.search(sent):
            continue
        score = len(_ACTION_SIGNAL.findall(sent))
        # Prefer English sentences when body is mixed
        if detect_lang(sent) == "en":
            score += 2
        scored.append((score, sent))

    # Pick top 2 sentences
    scored.sort(key=lambda x: -x[0])
    top = [s for _, s in scored[:2]]

    if translated_hints:
        result = " ".join(translated_hints)
        if top:
            result += " — " + " ".join(top)
    elif top:
        result = " ".join(top)
    else:
        # Fallback: first 200 chars
        result = body[:200].replace("\n", " ")

    # Trim and clean
    result = re.sub(r"\s{2,}", " ", result).strip()
    if len(result) > 400:
        result = result[:397] + "..."
    return result


# ── Email filter ───────────────────────────────────────────────────────────────

def should_skip(mail) -> bool:
    try:
        sender_email = safe_str(mail.SenderEmailAddress).lower()
        sender_name  = safe_str(mail.SenderName).lower()
        if RAPHAEL_EMAIL in sender_email:
            return True
        if sender_email in SKIP_SENDER_EMAILS:
            return True
        for domain in SKIP_SENDER_DOMAINS:
            if sender_email.endswith(f"@{domain}"):
                return True
        for kw in SKIP_SENDER_KEYWORDS:
            if kw in sender_name:
                return True
        return False
    except Exception:
        return False


# ── Row builder ────────────────────────────────────────────────────────────────

def build_row(mail, client_name: str) -> dict:
    received   = mail.ReceivedTime
    date_str   = received.strftime("%Y-%m-%d")
    body_raw   = safe_str(mail.Body)
    body_clean = clean_body(body_raw)
    subject    = safe_str(mail.Subject)
    from_name  = safe_str(mail.SenderName)
    from_email = safe_str(mail.SenderEmailAddress)

    to_str = ""
    try:
        recips   = mail.Recipients
        to_names = [recips.Item(i).Name for i in range(1, recips.Count + 1)
                    if recips.Item(i).Type == 1]
        to_str   = "; ".join(to_names)
    except Exception:
        pass

    meta     = parse_metadata(subject, body_clean)
    pm       = parse_pm(body_clean, to_str + " " + from_name)
    reply_fw, proj_code, batch_file, topic = split_subject(subject)
    summary  = summarize(body_clean, meta["Task Type"], meta["Provider"], meta["Deadline"])

    if not meta["Provider"] and from_name.lower() in QUOTED_PMS:
        meta["Provider"] = from_name

    if not meta["Project"] and proj_code:
        meta["Project"] = proj_code

    from_display = f"{from_name} <{from_email}>" if "@" in from_email else from_name

    if not is_project_related(subject, body_clean, meta["Task Type"]):
        return None

    return {
        "Date Received":  date_str,
        "Client":         client_name,
        "Project":        meta["Project"],
        "PM Responsible": pm,
        "From":           from_display,
        "Task Type":      meta["Task Type"],
        "Provider":       meta["Provider"],
        "Deadline":       meta["Deadline"],
        "Word Count":     meta["Word Count"],
        "Reply/FW":       reply_fw,
        "Project Code":   proj_code,
        "Batch / File":   batch_file,
        "Subject Topic":  topic,
        "Summary":        summary,
        "Entry ID":       safe_str(mail.EntryID),
        "_is_translation": meta["Task Type"] in {
            "Translation", "Proofreading", "Review", "LQA", "Post-Editing", "Delivery"
        },
    }


# ── Excel helpers ──────────────────────────────────────────────────────────────

def load_or_create_workbook():
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        header    = [cell.value for cell in ws[1]]
        eid_col   = header.index("Entry ID") + 1 if "Entry ID" in header else None
        existing  = set()
        if eid_col:
            for row in ws.iter_rows(min_row=2, min_col=eid_col, max_col=eid_col, values_only=True):
                if row[0]:
                    existing.add(row[0])
        return wb, existing
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title        = "Email Tracker"
        ws.freeze_panes = "A2"
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.column_dimensions[get_column_letter(len(COLUMNS))].hidden = True
        return wb, set()


def apply_row(ws, row_idx: int, row_data: dict):
    is_trans = row_data.get("_is_translation", False)
    fill     = (FILL_TRANS_A if row_idx % 2 == 0 else FILL_TRANS_B) if is_trans \
               else (FILL_GEN_A if row_idx % 2 == 0 else FILL_GEN_B)
    wc_col   = COLUMNS.index("Word Count") + 1

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        if col_name == "Entry ID":
            ws.cell(row=row_idx, column=col_idx, value=row_data.get("Entry ID", ""))
            continue
        cell           = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
        cell.fill      = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col_name == "Summary"))
        cell.border    = BORDER
        if col_idx == wc_col and is_trans and not row_data.get("Word Count"):
            cell.fill  = FILL_WC_MISSING
            cell.font  = FONT_WC_MISSING
            cell.value = "—"


def finalise(ws):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 15)
    ws.column_dimensions[get_column_letter(len(COLUMNS))].hidden = True
    last_data_col = get_column_letter(len(COLUMNS) - 1)
    ws.auto_filter.ref = f"A1:{last_data_col}1"


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Connecting to Outlook...")
    try:
        outlook     = win32com.client.Dispatch("Outlook.Application")
        ns          = outlook.GetNamespace("MAPI")
    except Exception as e:
        sys.exit(f"Could not connect to Outlook: {e}")

    terra_store = None
    for i in range(1, ns.Stores.Count + 1):
        store = ns.Stores.Item(i)
        if "terrateamup" in store.DisplayName.lower():
            terra_store = store
            break
    if not terra_store:
        sys.exit("Could not find terrateamup.com mailbox in Outlook.")

    root = terra_store.GetRootFolder()

    # Collect client folders: root-level + subfolders of Inbox
    client_folders = {}
    for i in range(1, root.Folders.Count + 1):
        f = root.Folders.Item(i)
        name_lower = f.Name.lower()
        if name_lower in SKIP_FOLDERS:
            continue
        if name_lower in ("caixa de entrada", "inbox"):
            # scan its subfolders as client folders
            for j in range(1, f.Folders.Count + 1):
                sub = f.Folders.Item(j)
                if sub.Name.lower() not in SKIP_FOLDERS:
                    client_folders[sub.Name] = sub
        else:
            client_folders[f.Name] = f

    if not client_folders:
        sys.exit("No client folders found.")

    print(f"Folders: {', '.join(client_folders)}")
    wb, existing_ids = load_or_create_workbook()
    print(f"Existing tracked: {len(existing_ids)}")

    new_rows, scanned, skipped = [], 0, 0

    for client_name, folder in client_folders.items():
        print(f"  Scanning {client_name}: {folder.Items.Count} items...")
        for i in range(1, folder.Items.Count + 1):
            try:
                item = folder.Items.Item(i)
                if item.Class != 43:
                    continue
                scanned += 1
                eid = safe_str(item.EntryID)
                if eid in existing_ids:
                    continue
                if should_skip(item):
                    skipped += 1
                    existing_ids.add(eid)
                    continue
                row = build_row(item, client_name)
                existing_ids.add(eid)
                if row:
                    new_rows.append(row)
            except Exception as ex:
                print(f"    Warning item {i} in {client_name}: {ex}")

    print(f"\nScanned {scanned} | Skipped: {skipped} | New rows: {len(new_rows)}")

    if new_rows:
        new_rows.sort(key=lambda r: r["Date Received"])
        ws       = wb.active
        next_row = ws.max_row + 1
        for row_data in new_rows:
            apply_row(ws, next_row, row_data)
            next_row += 1
        finalise(ws)
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(EXCEL_PATH)
        print(f"Saved {len(new_rows)} rows to {EXCEL_PATH}")
    else:
        print("No new emails. Tracker is up to date.")


if __name__ == "__main__":
    main()
